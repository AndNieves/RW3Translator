createUniqueString = '''
--Create Unique String {str_key}
EXEC dbo.CreateUniqueString '{str_key}'
'''

createTranslation = '''
--Create {lang} Translation
EXEC dbo.CreateTranslation '{str_key}', '{lang}', N'{lang_text}'
'''

createFileString = '''
--Associating key str_key with page {page}
EXEC dbo.CreateFileString '{str_key}','{page}'
'''


languages = ['Chinese',
             'ChineseTW',
             'French',
             'German',
             'Italian',
             'Japanese',
             'Korean',
             'Spanish',
             'English',
             'Russian',
             'Catalan']



from openpyxl import load_workbook
wb = load_workbook('./LRW_Translations.xlsx')
ws = wb['original']
first = True

#print ('SET QUOTED_IDENTIFIER OFF;')
for row in ws.iter_rows():
    if first:
        firstRow = row;
        first = False;
        continue;
    key = row[1].value
    if key is not None:
        print (createUniqueString.format(str_key=key.strip()))
        pages = row[2].value.split(',')
        for page in pages:
            print (createFileString.format(str_key=key,page=page.strip()))
        for i in range(0, len(languages)):
            currentIndex = i + 3
            languageName = firstRow[currentIndex].value
            value = row[currentIndex].value
            if value is not None:
                valueEscaped = value.replace('\'', '\'\'\'\'')
                print(createTranslation.format(str_key=key,lang=languageName,lang_text=valueEscaped))
        print('GO')
#print ('SET QUOTED_IDENTIFIER ON;')